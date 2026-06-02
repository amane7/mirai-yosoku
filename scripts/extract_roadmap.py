#!/usr/bin/env python3
"""
Future Road Map (xlsx) -> structured TS extraction.

Emits two files:
  src/data/roadmap-data.ts  - full browsable dataset (4 axes x 7 eras keywords + Tech Effect + PEST)
  src/data/roadmap.ts       - a runnable Project (eras->time, axes->categories,
                              keyword groups->sub-factors, Tech Effect->edges, PEST->scenarios)

要素分解 (element decomposition) is first-class:
  era x axis  ->  Factor  (28 factors = 7 eras x 4 axes)
  each keyword inside that cell -> SubFactor of its Factor
"""
import json
import re
import openpyxl

SRC = "/tmp/roadmap/roadmap.xlsx"
OUT_DIR = "/home/user/webapp/src/data"

# Era blocks discovered from column A markers
ERA_STARTS = {
    "2020-2024": 2,
    "2025-2029": 53,
    "2030-2034": 104,
    "2035-2039": 155,
    "2040-2044": 206,
    "2045-2049": 257,
    "2050-2054": 308,
}
ERA_ORDER = list(ERA_STARTS.keys())

# axis column groups on the main sheet: (No, keyword, summary)
AXES = {
    "technology": {"no": 2, "kw": 3, "sum": 4, "label": "テクノロジー", "category": "Technology"},
    "market": {"no": 6, "kw": 7, "sum": 8, "label": "マーケット", "category": "Market"},
    "literacy": {"no": 10, "kw": 11, "sum": 12, "label": "リテラシー", "category": "People"},
    "culture": {"no": 14, "kw": 15, "sum": 16, "label": "カルチャー", "category": "Culture"},
}
AXIS_ORDER = ["technology", "market", "literacy", "culture"]


def cell(ws, r, c):
    v = ws.cell(row=r, column=c).value
    if v is None:
        return ""
    return str(v).strip()


def era_year(era_key):
    """Mid-year of an era like '2030-2034' -> (start,mid,end)."""
    a, b = era_key.split("-")
    a, b = int(a), int(b)
    return a, (a + b) // 2, b


def extract_keywords(wb):
    ws = wb["未来ロードマップキーワード第1弾"]
    eras = []
    for ei, era_key in enumerate(ERA_ORDER):
        start = ERA_STARTS[era_key]
        end = ERA_STARTS[ERA_ORDER[ei + 1]] - 1 if ei + 1 < len(ERA_ORDER) else ws.max_row
        axes_out = {}
        for axis in AXIS_ORDER:
            ax = AXES[axis]
            items = []
            for r in range(start, end + 1):
                kw = cell(ws, r, ax["kw"])
                if not kw:
                    continue
                no = cell(ws, r, ax["no"])
                try:
                    no = int(float(no))
                except (ValueError, TypeError):
                    no = len(items) + 1
                items.append({
                    "no": no,
                    "name": kw,
                    "summary": cell(ws, r, ax["sum"]),
                })
            axes_out[axis] = items
        eras.append({"era": era_key, "axes": axes_out})
    return eras


def extract_tech_effect(wb):
    ws = wb["Tech Effect"]
    # header at row 1: A年代 B No C keyword D summary E +market F -market G +impact H -impact
    out = {}  # (era, no) -> dict
    for r in range(2, ws.max_row + 1):
        kw = cell(ws, r, 3)
        if not kw:
            continue
        era = cell(ws, r, 1)
        no = cell(ws, r, 2)
        try:
            no = int(float(no))
        except (ValueError, TypeError):
            continue
        out[f"{era}#{no}"] = {
            "era": era,
            "no": no,
            "name": kw,
            "posMarket": cell(ws, r, 5),
            "negMarket": cell(ws, r, 6),
            "posImpact": cell(ws, r, 7),
            "negImpact": cell(ws, r, 8),
        }
    return out


def split_industries(text):
    """Parse '・SaaS開発（高） ・教育（中）' -> [{industry, level}]."""
    if not text:
        return []
    out = []
    # split on '・' bullets and newlines
    parts = re.split(r"[・\n]", text)
    for p in parts:
        p = p.strip()
        if not p:
            continue
        m = re.search(r"[（(]\s*([高中低])\s*[)）]", p)
        level = m.group(1) if m else ""
        name = re.sub(r"[（(].*?[)）]", "", p).strip()
        if name:
            out.append({"industry": name, "level": level})
    return out


def extract_pest(wb):
    """Use PEST3 (year x P/E/S/T keyword:desc) as the primary, structured source."""
    ws = wb["PEST3"]
    # header at row 2: B年代 C P D E E S F T
    pest = []
    for r in range(3, 12):
        year = cell(ws, r, 2)
        if not year:
            continue
        def parse_cell(c):
            raw = cell(ws, r, c)
            items = []
            for seg in raw.split("/"):
                seg = seg.strip()
                if not seg:
                    continue
                if ":" in seg or "：" in seg:
                    parts2 = re.split(r"[:：]", seg, maxsplit=1)
                    k = parts2[0]
                    v = parts2[1] if len(parts2) > 1 else ""
                    items.append({"keyword": k.strip(), "desc": v.strip()})
                else:
                    items.append({"keyword": seg, "desc": ""})
            return items
        pest.append({
            "year": year,
            "politics": parse_cell(3),
            "economy": parse_cell(4),
            "society": parse_cell(5),
            "technology": parse_cell(6),
        })
    return pest


def main():
    wb = openpyxl.load_workbook(SRC, data_only=True)
    eras = extract_keywords(wb)
    tech_effect = extract_tech_effect(wb)
    pest = extract_pest(wb)

    # attach tech effect (+ parsed industries) to technology keywords
    for blk in eras:
        for it in blk["axes"]["technology"]:
            te = tech_effect.get(f"{blk['era']}#{it['no']}")
            if te:
                it["posMarket"] = te["posMarket"]
                it["negMarket"] = te["negMarket"]
                it["posIndustries"] = split_industries(te["posImpact"])
                it["negIndustries"] = split_industries(te["negImpact"])

    total_kw = sum(len(blk["axes"][a]) for blk in eras for a in AXIS_ORDER)
    print(f"Extracted: {len(eras)} eras, {total_kw} keywords, "
          f"{len(tech_effect)} tech-effect rows, {len(pest)} PEST years")

    payload = {"eras": eras, "pest": pest}
    with open("/tmp/roadmap/roadmap.json", "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=1)
    print("Wrote /tmp/roadmap/roadmap.json")


if __name__ == "__main__":
    main()
